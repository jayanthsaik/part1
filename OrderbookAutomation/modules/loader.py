from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict

import openpyxl
import pandas as pd

from config import AppConfig
from modules.source_discovery import ResolvedSource, discover_sources


@dataclass(frozen=True)
class WorksheetData:
    """Container for worksheet payload and validation metadata."""

    workbook_name: str
    sheet_name: str
    header_row_index: int
    dataframe: pd.DataFrame
    hidden_columns: tuple[str, ...]
    merged_header_ranges: tuple[str, ...]
    raw_headers: tuple[str, ...]
    duplicate_raw_headers: tuple[str, ...]
    blank_raw_headers: tuple[str, ...]


@dataclass(frozen=True)
class WorkbookData:
    """Container for workbook payload."""

    workbook_key: str
    workbook_name: str
    workbook_path: Path
    worksheets: Dict[str, WorksheetData]


def _read_hidden_and_merged(workbook_path: Path, sheet_name: str, dataframe_columns: list[str]) -> tuple[tuple[str, ...], tuple[str, ...]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    worksheet = workbook[sheet_name]

    hidden_columns: list[str] = []
    for index, column_name in enumerate(dataframe_columns, start=1):
        column_letter = openpyxl.utils.get_column_letter(index)
        if worksheet.column_dimensions[column_letter].hidden:
            hidden_columns.append(column_name)

    merged_header_ranges = tuple(str(cell_range) for cell_range in worksheet.merged_cells.ranges)
    return tuple(hidden_columns), merged_header_ranges


def _extract_header_metadata(workbook_path: Path, sheet_name: str, header_row_index: int) -> tuple[tuple[str, ...], tuple[str, ...], tuple[str, ...]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    worksheet = workbook[sheet_name]

    header_row_number = header_row_index + 1
    row_values: list[str] = []
    for row in worksheet.iter_rows(min_row=header_row_number, max_row=header_row_number, values_only=True):
        row_values = ["" if value is None else str(value) for value in row]
        break

    if not row_values:
        return tuple(), tuple(), tuple()

    duplicate_headers = sorted({value for value in row_values if value.strip() and row_values.count(value) > 1})
    blank_headers = tuple(value for value in row_values if value.strip() == "")
    return tuple(row_values), tuple(duplicate_headers), blank_headers


def load_workbooks(config: AppConfig, logger) -> Dict[str, WorkbookData]:
    """Discover and load every configured logical source workbook.

    File and worksheet identification is handled entirely by
    ``modules/source_discovery.py`` based on business header content, not by
    fixed filenames or worksheet position. This keeps the pipeline working
    even when input files are renamed or re-dated (e.g. daily inventory
    exports).
    """
    loaded: Dict[str, WorkbookData] = {}

    resolved_sources: Dict[str, ResolvedSource] = discover_sources(config.source_definitions, config.input_dir, logger)

    for logical_name, resolved in resolved_sources.items():
        workbook_path = resolved.file_path
        sheet_name = resolved.sheet_name
        header_row = resolved.header_row_index

        try:
            dataframe = pd.read_excel(workbook_path, sheet_name=sheet_name, header=header_row)
        except Exception as exc:
            raise ValueError(
                f"Failed to read workbook '{workbook_path.name}' sheet '{sheet_name}' with header row {header_row}: {exc}"
            ) from exc

        hidden_columns, merged_header_ranges = _read_hidden_and_merged(workbook_path, sheet_name, list(dataframe.columns))
        raw_headers, duplicate_raw_headers, blank_raw_headers = _extract_header_metadata(workbook_path, sheet_name, header_row)

        worksheets: Dict[str, WorksheetData] = {
            sheet_name: WorksheetData(
                workbook_name=workbook_path.name,
                sheet_name=sheet_name,
                header_row_index=header_row,
                dataframe=dataframe,
                hidden_columns=hidden_columns,
                merged_header_ranges=merged_header_ranges,
                raw_headers=raw_headers,
                duplicate_raw_headers=duplicate_raw_headers,
                blank_raw_headers=blank_raw_headers,
            )
        }
        logger.info("Loaded workbook=%s sheet=%s rows=%s cols=%s", workbook_path.name, sheet_name, len(dataframe), len(dataframe.columns))

        loaded[logical_name] = WorkbookData(
            workbook_key=logical_name,
            workbook_name=workbook_path.name,
            workbook_path=workbook_path,
            worksheets=worksheets,
        )

    return loaded
