from __future__ import annotations

import logging
import re
import sys
from datetime import date, datetime
from pathlib import Path
from time import perf_counter
from typing import Iterable

import openpyxl
import pandas as pd


def ensure_directory(path: Path) -> Path:
    """Create the directory if it does not exist and return it."""
    path.mkdir(parents=True, exist_ok=True)
    return path


def get_application_base_dir() -> Path:
    """Return the folder that should be treated as the application's home directory.

    When running as a normal Python script, this is the folder containing the
    project source (so paths stay stable regardless of the current working
    directory). When running as a frozen executable (e.g. built with
    PyInstaller), ``sys.executable`` points at the .exe itself, so we use its
    parent folder instead. This lets the business drop input Excel files next
    to the .exe and have them picked up automatically, without ever touching
    source code or hardcoded paths.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent.parent


def _is_blank_value(value: object) -> bool:
    return clean_string_value(value) is None


def _header_row_score(row: Iterable[object]) -> int:
    """Score a worksheet row on how likely it is to be the real header row."""
    values = [value for value in row if not _is_blank_value(value)]
    if not values:
        return -(10**9)
    string_count = sum(isinstance(value, str) for value in values)
    date_count = sum(isinstance(value, (datetime, date)) for value in values)
    numeric_count = sum(isinstance(value, (int, float)) and not isinstance(value, bool) for value in values)
    return string_count * 3 + date_count * 2 + len(values) - numeric_count


def detect_header_row(workbook_path: Path, sheet_name: str, max_rows_to_scan: int = 25) -> int:
    """Automatically detect the most likely header row index (0-based) for a worksheet.

    This removes the need to hardcode a header row index per workbook/sheet in
    config.py: each workbook is scanned and the row that looks most like a
    header (mostly text, few numeric-only cells) is selected automatically.
    """
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet_name]
        best_index = 0
        best_score = -(10**9)
        max_row = min(worksheet.max_row or 1, max_rows_to_scan)
        for index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_row, values_only=True)):
            score = _header_row_score(row)
            if score > best_score:
                best_score = score
                best_index = index
        return best_index
    finally:
        workbook.close()


def clean_string_value(value: object) -> str | None:
    """Return a cleaned string value while keeping original letter casing."""
    if pd.isna(value):
        return None
    text = str(value)
    text = text.replace("\t", " ").replace("\r", " ").replace("\n", " ")
    text = "".join(char for char in text if char.isprintable())
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def is_blank_hold_code(value: object) -> bool:
    """Return True when a Hold Codes value must be treated as blank.

    Centralized helper for the UPS Inventory Hold Codes eligibility rule.
    Treats as blank: actual Excel blank/None/NaN, empty string, and
    whitespace-only strings (including tabs/newlines with no real content).
    Any other, non-blank text (e.g. "HOLD", "QC", "DAMAGED") is NOT blank.
    """
    return clean_string_value(value) is None


def normalize_text_key(value: object) -> str | None:
    """Normalize a text key for resilient, case-insensitive lookup joins."""
    cleaned = clean_string_value(value)
    if cleaned is None:
        return None
    return cleaned.upper()


def normalize_identifier_key(value: object) -> str | None:
    """Normalize identifier-like keys without forcing uppercase."""
    return clean_string_value(value)


def normalize_ndc_key(value: object) -> str | None:
    """Normalize an NDC value for join/matching purposes only.

    This function is used exclusively to build temporary comparison keys for
    NDC-based joins. It never modifies the original business "NDC Code"
    column value stored elsewhere in a dataframe.

    Handles:
    - NaN / None -> None
    - Safe conversion to string
    - Leading/trailing whitespace trimming
    - Excel-introduced ".0" float artifacts (e.g. "64380201.0" -> "64380201")
    - NDC separators such as "-" (e.g. "64380-201-01" -> "6438020101")
    - Other non-numeric formatting characters (spaces, slashes, dots that are
      not part of a genuine decimal value)
    - Leading zeros are preserved because no numeric int() conversion is
      performed; only non-digit characters are stripped from the string.
    """
    if pd.isna(value):
        return None

    if isinstance(value, float):
        # Whole-number floats (typical of Excel-stored numeric NDC values,
        # e.g. 64380201.0) should not retain a trailing ".0".
        if value.is_integer():
            text = str(int(value))
        else:
            text = str(value)
    else:
        text = str(value)

    text = text.strip()
    if not text:
        return None

    # Drop a trailing Excel float artifact such as "64380201.0" that may have
    # survived string coercion upstream (e.g. object dtype containing str
    # already formatted like "64380201.0").
    text = re.sub(r"\.0+$", "", text)

    # Strip everything that is not a digit (dashes, spaces, slashes, etc.)
    # while preserving digit order, and therefore any leading zeros.
    digits_only = re.sub(r"\D", "", text)

    return digits_only or None


def strip_trailing_zero(value: object) -> object:
    """Convert numeric-looking strings ending in .0 into integer-like strings."""
    cleaned = clean_string_value(value)
    if cleaned is None:
        return None
    if re.fullmatch(r"[+-]?\d+\.0+", cleaned):
        return cleaned.split(".", maxsplit=1)[0]
    return cleaned


def coerce_numeric_column(series: pd.Series) -> pd.Series:
    """Coerce a Series to numeric values where possible."""
    return pd.to_numeric(series, errors="coerce")


def coerce_string_column(series: pd.Series) -> pd.Series:
    """Coerce a Series to pandas string dtype with empty values as NA."""
    coerced = series.astype("string")
    return coerced.replace("", pd.NA)


def setup_logging(log_dir: Path, log_name: str = "phase1.log") -> logging.Logger:
    """Configure logger for Phase 1 ingestion."""
    ensure_directory(log_dir)
    logger = logging.getLogger("orderbook_automation_phase1")
    logger.setLevel(logging.INFO)
    logger.propagate = False

    if not logger.handlers:
        formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
        file_handler = logging.FileHandler(log_dir / log_name, encoding="utf-8")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

        stream_handler = logging.StreamHandler()
        stream_handler.setFormatter(formatter)
        logger.addHandler(stream_handler)

    return logger


class Timer:
    """Simple context timer to track execution duration."""


    def __init__(self) -> None:
        self.start = 0.0
        self.end = 0.0

    def __enter__(self) -> "Timer":
        self.start = perf_counter()
        return self


    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.end = perf_counter()


    @property
    def elapsed_seconds(self) -> float:
        if self.start == 0.0:
            return 0.0
        if self.end == 0.0:
            return max(perf_counter() - self.start, 0.0)
        return max(self.end - self.start, 0.0)
