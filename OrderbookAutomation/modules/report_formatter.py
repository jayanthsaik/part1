from __future__ import annotations

import os
from pathlib import Path
from typing import Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from modules.business_rules import LOW_UPS_INVENTORY_THRESHOLD

# Documented SOP colors (Part E).
FILL_CONTROLLED_PRODUCT_PINK = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
FILL_PRICE_ISSUE_ORANGE = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
FILL_LOW_INVENTORY_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_CANCEL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
FILL_HOLD_BLUE = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

_MOQ_TOLERANCE = 1e-9


def _is_moq_issue(sales_order_qty: object, pack_size: object) -> bool:
    """Return True when Sales Order Qty is not an exact multiple of Pack
    Size (MOQ), per the documented manual process: "divide Sales Order
    Quantity by Pack Size (MOQ). Any lines with a number after the decimal
    point are not in MOQ."

    Computed directly from the final Orderbook row's own worksheet cell
    values (never from the optional Phase 2 sales_summary source). Safe
    for blank/zero/non-numeric inputs (returns False, never raises). Uses
    a small floating-point tolerance so a mathematically exact multiple
    (e.g. 72 / 24 = 3.0) is never falsely flagged due to float noise.
    """
    try:
        if sales_order_qty is None or pack_size is None:
            return False
        qty = float(sales_order_qty)
        size = float(pack_size)
    except (TypeError, ValueError):
        return False

    if pd.isna(qty) or pd.isna(size):
        return False
    if size == 0:
        return False

    remainder = qty % size
    # Treat remainders within tolerance of 0 OR of the full pack size
    # (floating point can produce e.g. 23.999999999997 instead of 0) as
    # an exact multiple, not an MOQ issue.
    if remainder <= _MOQ_TOLERANCE or (size - remainder) <= _MOQ_TOLERANCE:
        return False
    return True


def _sanitize_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy of ``df`` with pandas NA/NaT/NaN and numpy scalar types replaced by
    plain Python values that openpyxl can write to a cell.

    openpyxl cannot serialize ``pd.NA`` (pandas' nullable-dtype missing marker) or numpy
    scalar types directly; converting to ``object`` dtype and replacing missing markers
    with ``None`` avoids "Cannot convert <NA> to Excel" errors while preserving actual
    business values (0 stays 0, empty string stays empty string, etc.).
    """
    safe_df = df.copy()
    for column in safe_df.columns:
        safe_df[column] = safe_df[column].astype(object).where(safe_df[column].notna(), None)
        safe_df[column] = safe_df[column].map(_to_excel_scalar)
    return safe_df


def _to_excel_scalar(value):
    if value is None or value is pd.NA:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if pd.api.types.is_scalar(value) and pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp,)):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        try:
            return value.item()
        except (ValueError, AttributeError):
            return value
    return value


def write_dataframe_sheet(
    workbook: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    freeze_header: bool = True,
    autofilter: bool = True,
) -> Worksheet:
    """Write ``df`` as values (never formulas) into a new worksheet with standard formatting."""
    worksheet = workbook.create_sheet(title=sheet_name)
    safe_df = _sanitize_for_excel(df)
    for row in dataframe_to_rows(safe_df, index=False, header=True):
        worksheet.append(row)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    if freeze_header:
        worksheet.freeze_panes = "A2"
    if autofilter and worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    _autosize_columns(worksheet)
    return worksheet


def _autosize_columns(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 10), 55)


def apply_low_ups_inventory_formatting(
    worksheet: Worksheet,
    *,
    ups_inventory_column: str = "Max of UPS Inventory",
    threshold: float = LOW_UPS_INVENTORY_THRESHOLD,
) -> int:
    """Fill the ``ups_inventory_column`` cell yellow where its value is below
    ``threshold``, and return the number of cells highlighted.

    This is the ONLY place the Low UPS Inventory rule is applied; it targets
    the aggregated Summary sheet's "Max of UPS Inventory" column and never
    the client-facing Orderbook sheet's "UPS Inventory" column. Blank and
    non-numeric values are skipped safely (no highlight, no error).
    """
    header_to_column_index = {str(cell.value): cell.column for cell in worksheet[1]}
    column_index = header_to_column_index.get(ups_inventory_column)
    if column_index is None:
        return 0

    highlighted = 0
    for row_number in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_number, column=column_index)
        value = cell.value
        if value is None or isinstance(value, bool):
            continue
        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            continue
        if pd.isna(numeric_value):
            continue
        if numeric_value < threshold:
            cell.fill = FILL_LOW_INVENTORY_YELLOW
            highlighted += 1

    return highlighted


def apply_business_rule_formatting(
    worksheet: Worksheet,
    df: pd.DataFrame,
    *,
    controlled_product_column: str = "Controlled_Product",
    price_issue_column: str = "Price_Issue",
    low_inventory_column: str = "Low_UPS_Inventory",
    price_columns: Sequence[str] = ("Unit Price", "WAC/BG price in EDI"),
    cancel_column: str | None = "Action",
    hold_column: str | None = "Action",
    sales_order_qty_column: str = "Sales Order Qty",
    pack_size_column: str = "Pack Size (MOQ)",
    ups_inventory_column: str = "UPS Inventory",
) -> None:
    """Apply the documented Part E color-coding rules to a written worksheet.

    Formatting is applied strictly after business-rule flags have already
    been computed (Part B); this function never computes new business
    conditions, it only reads existing flag values -- EXCEPT for the MOQ
    Issue highlight below, which is computed directly from the written
    worksheet's ``Sales Order Qty`` / ``Pack Size (MOQ)`` cells (a
    final-Orderbook-only formatting enhancement, independent of the
    optional Phase 2 sales_summary source).

    ``df`` is the internal working dataframe (e.g. ``enriched_df``) that
    still contains the internal flag columns (Controlled_Product,
    Price_Issue, Low_UPS_Inventory), even when those columns have been
    excluded from the written client-facing ``worksheet`` (per the
    authoritative reference schema). Flags are read positionally from
    ``df`` (row order/grain is preserved 1:1 between ``df`` and
    ``worksheet``, since the client-facing Orderbook performs no
    aggregation or row reduction); any business columns still present on
    the worksheet itself (e.g. "Action", price columns) are read directly
    from the worksheet as before.

    Fill precedence (applied in this order, later overwrites earlier):

    1. Controlled Product -> WHOLE ROW pink.
    2. Price Issue -> the specific price cells orange (overrides pink on
       those cells only).
    3. MOQ Issue -> the "Sales Order Qty" CELL ONLY yellow (overrides pink
       on that cell only, same "cell overrides row" pattern as the orange
       price cells). The MOQ rule NEVER fills the whole row.
    4. Action = Cancel -> WHOLE ROW red; Action = Hold -> WHOLE ROW blue.
       Applied last, so Cancel/Hold take FULL precedence over the pink row
       fill and over the orange price / yellow MOQ cell highlights.

    MOQ Issue rule (per the documented manual process): when
    ``Sales Order Qty`` is not an exact multiple of ``Pack Size (MOQ)``
    (computed directly from those two worksheet cells, header-driven --
    never a hardcoded column letter), the ``Sales Order Qty`` cell is
    filled yellow. Blank/zero/non-numeric Pack Size or Sales Order Qty are
    handled safely: no highlight, no error.

    NOTE: the Low UPS Inventory yellow highlight is intentionally NOT
    applied here. That rule lives on the Summary sheet's
    "Max of UPS Inventory" column (see
    ``apply_low_ups_inventory_formatting``), so the client-facing
    POB.xlsx "UPS Inventory" cells are never highlighted.
    """
    header_to_column_index = {str(cell.value): cell.column for cell in worksheet[1]}

    controlled_series = df[controlled_product_column] if controlled_product_column in df.columns else None
    price_issue_series = df[price_issue_column] if price_issue_column in df.columns else None

    price_col_indexes = [header_to_column_index[column] for column in price_columns if column in header_to_column_index]
    cancel_idx = header_to_column_index.get(cancel_column) if cancel_column else None
    hold_idx = header_to_column_index.get(hold_column) if hold_column else None
    sales_order_qty_idx = header_to_column_index.get(sales_order_qty_column)
    pack_size_idx = header_to_column_index.get(pack_size_column)

    df_row_values = df.reset_index(drop=True)

    for row_number in range(2, worksheet.max_row + 1):
        df_index = row_number - 2  # worksheet row 2 == df row 0 (header occupies row 1)
        is_controlled = (
            controlled_series is not None
            and df_index < len(df_row_values)
            and bool(df_row_values[controlled_product_column].iloc[df_index]) is True
        )
        is_price_issue = (
            price_issue_series is not None
            and df_index < len(df_row_values)
            and bool(df_row_values[price_issue_column].iloc[df_index]) is True
        )

        # MOQ Issue state computed independently, BEFORE any fill decision,
        # directly from this row's own worksheet cells.
        is_moq_issue = False
        if sales_order_qty_idx is not None and pack_size_idx is not None:
            sales_order_qty_value = worksheet.cell(row=row_number, column=sales_order_qty_idx).value
            pack_size_value = worksheet.cell(row=row_number, column=pack_size_idx).value
            is_moq_issue = _is_moq_issue(sales_order_qty_value, pack_size_value)

        # Whole-row fill for Controlled Product only. MOQ no longer fills the
        # whole row -- it is a single-cell highlight applied further below.
        if is_controlled:
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_number, column=col).fill = FILL_CONTROLLED_PRODUCT_PINK

        if is_price_issue:
            for column_index in price_col_indexes:
                worksheet.cell(row=row_number, column=column_index).fill = FILL_PRICE_ISSUE_ORANGE

        # MOQ Issue: highlight ONLY the "Sales Order Qty" cell yellow. This
        # follows the same "cell overrides row" pattern as the orange price
        # cells, so the yellow remains visible inside an otherwise-pink
        # Controlled Product row.
        if is_moq_issue and sales_order_qty_idx is not None:
            worksheet.cell(row=row_number, column=sales_order_qty_idx).fill = FILL_LOW_INVENTORY_YELLOW

        # Action-based colors (Cancel = red, Hold = blue). These read an
        # existing "Action" business field if present; no automatic
        # business rule creates or infers this value. Applied LAST so a
        # Cancel/Hold row takes full precedence over the pink row fill and
        # over the orange price / yellow MOQ cell highlights.
        if cancel_idx is not None:
            action_value = str(worksheet.cell(row=row_number, column=cancel_idx).value or "").strip().lower()
            if action_value == "cancel":
                for column_index in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=row_number, column=column_index).fill = FILL_CANCEL_RED
            elif hold_idx is not None and action_value == "hold":
                for column_index in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=row_number, column=column_index).fill = FILL_HOLD_BLUE


def save_workbook(workbook: Workbook, output_path: Path) -> None:
    """Save ``workbook`` ATOMICALLY to ``output_path``.

    The workbook is first written to a temporary file in the same folder and
    then atomically moved into place, so a failure mid-write can never leave
    a partial or corrupt output file behind. The temp artifact is always
    cleaned up on error.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        default_sheet = workbook["Sheet"]
        if default_sheet.max_row == 1 and default_sheet.max_column == 1 and default_sheet["A1"].value is None:
            workbook.remove(default_sheet)

    temp_path = output_path.with_name(f"{output_path.stem}.__writing__{output_path.suffix}")
    try:
        workbook.save(temp_path)
        os.replace(temp_path, output_path)
    except BaseException:
        # Never leave the partial temp artifact behind on failure.
        try:
            if temp_path.exists():
                temp_path.unlink()
        except OSError:
            pass
        raise
