from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Sequence

import openpyxl

from modules.source_registry import SourceDefinition
from modules.utils import detect_header_row


class SourceDiscoveryError(Exception):
    """Base class for business-friendly source discovery failures."""


class SourceNotFoundError(SourceDiscoveryError):
    """Raised when no file/sheet in the input folder matches a required source."""

    def __init__(self, source: SourceDefinition, input_dir: Path) -> None:
        message = (
            f"Could not find the '{source.display_name}' workbook in the input folder "
            f"({input_dir}).\n"
            f"Expected a worksheet containing these required columns: {list(source.required_headers)}.\n"
            f"Hint: {source.filename_hint or 'no additional hint available'}."
        )
        super().__init__(message)
        self.source = source
        self.input_dir = input_dir


class AmbiguousSourceError(SourceDiscoveryError):
    """Raised when more than one candidate file/sheet matches a required source."""

    def __init__(self, source: SourceDefinition, candidates: Sequence["SourceCandidate"]) -> None:
        candidate_list = "\n".join(f"  - {candidate.file_path.name} (worksheet '{candidate.sheet_name}')" for candidate in candidates)
        message = (
            f"Multiple files/worksheets match the '{source.display_name}' source and the "
            "correct one could not be determined automatically:\n"
            f"{candidate_list}\n"
            "Please leave only the correct/current workbook in the input folder and re-run."
        )
        super().__init__(message)
        self.source = source
        self.candidates = list(candidates)


@dataclass(frozen=True)
class SourceCandidate:
    """A discovered file/worksheet pairing that matched a source definition."""

    file_path: Path
    sheet_name: str
    header_row_index: int
    matched_required_count: int
    is_preferred_sheet: bool


@dataclass(frozen=True)
class ResolvedSource:
    """Final resolved location for a logical source."""

    logical_name: str
    file_path: Path
    sheet_name: str
    header_row_index: int


def _read_sheet_headers(file_path: Path, sheet_name: str, header_row_index: int) -> List[str]:
    """Read the raw header row exactly as it appears in the workbook.

    Header text must be preserved verbatim (including any intentional
    leading/trailing spaces used in real business headers, e.g. " Total ")
    so that discovery matching against ``required_headers`` in
    ``source_registry.py`` is an exact match, not a normalized one.
    """
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet_name]
        header_row_number = header_row_index + 1
        row_values: List[str] = []
        for row in worksheet.iter_rows(min_row=header_row_number, max_row=header_row_number, values_only=True):
            row_values = ["" if value is None else str(value) for value in row]
            break
        return row_values
    finally:
        workbook.close()


def _sheet_has_data_rows(file_path: Path, sheet_name: str, header_row_index: int) -> bool:
    """Return True if the worksheet has at least one non-empty row after the header row.

    This is a generic, content-based check (not a filename check) used to
    exclude header-only template/reference worksheets from discovery -- any
    worksheet whose header row matches a source's required headers but which
    contains no actual data rows underneath cannot be a real business data
    source, regardless of what the file happens to be named.
    """
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet_name]
        first_data_row_number = header_row_index + 2
        for row in worksheet.iter_rows(min_row=first_data_row_number, values_only=True):
            if any(value is not None and str(value).strip() for value in row):
                return True
        return False
    finally:
        workbook.close()


def discover_sources(
    source_definitions: Dict[str, SourceDefinition],
    input_dir: Path,
    logger,
) -> Dict[str, ResolvedSource]:
    """Scan ``input_dir`` and resolve each logical source to a concrete file + worksheet.

    Discovery is based entirely on worksheet header content (Step 5/6 rules),
    never on filename pattern matching or worksheet position/index. Filenames
    may change freely (including date-stamped daily exports) without
    affecting discovery.
    """
    excel_files = sorted(input_dir.glob("*.xlsx"))
    # Files this application itself generates (see main.py / phase4_manager.py
    # output filenames) must never be treated as candidate input sources,
    # even if a stale/previous copy is accidentally left in the input
    # folder. These are outputs, not inputs, regardless of header shape.
    generated_output_filenames = {
        "workbook_profile.xlsx",
        "derived_data.xlsx",
        "business_master_data.xlsx",
        "sales_summary.xlsx",
        "pob.xlsx",
    }
    excel_files = [
        file_path
        for file_path in excel_files
        if file_path.name.strip().lower() not in generated_output_filenames
    ]
    if not excel_files:
        raise SourceNotFoundError(next(iter(source_definitions.values())), input_dir)

    # Pre-scan every workbook/sheet once, header row detected dynamically.
    workbook_sheet_headers: List[tuple[Path, str, int, List[str]]] = []
    for file_path in excel_files:
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001 - surfaced as business-friendly warning
            logger.warning("Skipping unreadable workbook '%s': %s", file_path.name, exc)
            continue
        try:
            sheet_names = workbook.sheetnames
        finally:
            workbook.close()

        for sheet_name in sheet_names:
            header_row_index = detect_header_row(file_path, sheet_name)
            headers = _read_sheet_headers(file_path, sheet_name, header_row_index)
            if not any(header.strip() for header in headers):
                continue
            # Skip header-only template/reference worksheets: a worksheet
            # whose header row matches but has no actual data rows beneath
            # it cannot be a real business data source. This is a
            # content-based check, never a filename check, so it applies
            # equally regardless of what the workbook happens to be named.
            if not _sheet_has_data_rows(file_path, sheet_name, header_row_index):
                logger.info(
                    "Discovery scan | file=%s | sheet=%s | skipped (no data rows beneath header)",
                    file_path.name,
                    sheet_name,
                )
                continue
            workbook_sheet_headers.append((file_path, sheet_name, header_row_index, headers))
            logger.info(
                "Discovery scan | file=%s | sheet=%s | header_row=%s | header_count=%s",
                file_path.name,
                sheet_name,
                header_row_index,
                len([h for h in headers if h.strip()]),
            )

    resolved: Dict[str, ResolvedSource] = {}

    for logical_name, source in source_definitions.items():
        candidates: List[SourceCandidate] = []

        for file_path, sheet_name, header_row_index, headers in workbook_sheet_headers:
            header_set = set(headers)
            matched_required = [header for header in source.required_headers if header in header_set]
            if len(matched_required) != len(source.required_headers):
                continue
            candidates.append(
                SourceCandidate(
                    file_path=file_path,
                    sheet_name=sheet_name,
                    header_row_index=header_row_index,
                    matched_required_count=len(matched_required),
                    is_preferred_sheet=sheet_name in source.preferred_sheet_names,
                )
            )

        if not candidates:
            if not source.mandatory:
                logger.warning(
                    "Optional source '%s' (%s) not found in input folder; skipping (not required for this workflow).",
                    logical_name,
                    source.display_name,
                )
                continue
            raise SourceNotFoundError(source, input_dir)

        if len(candidates) > 1:
            preferred = [candidate for candidate in candidates if candidate.is_preferred_sheet]
            # Distinct underlying files among all candidates (ignoring sheet duplicates
            # within the same file, which just means the sheet also matched elsewhere
            # in the same workbook and is not truly ambiguous at the file level).
            distinct_files = {candidate.file_path for candidate in candidates}

            if len(distinct_files) == 1:
                # Same file, multiple matching sheets: prefer the configured
                # preferred sheet name if exactly one such sheet exists.
                if len(preferred) == 1:
                    candidates = preferred
                elif len(preferred) == 0 and len(candidates) == 1:
                    pass
                else:
                    if not source.mandatory:
                        logger.warning(
                            "Optional source '%s' (%s) is ambiguous (multiple matching sheets); skipping (not required for this workflow).",
                            logical_name,
                            source.display_name,
                        )
                        continue
                    raise AmbiguousSourceError(source, candidates)
            else:
                # Multiple distinct files match: this is a genuine business
                # ambiguity (e.g. two inventory exports left in the folder).
                if len(preferred) == 1:
                    candidates = preferred
                elif source.filename_keywords:
                    # Last-resort tiebreaker: some workbooks are byte-identical
                    # in header shape (e.g. a header-only template/reference
                    # copy vs. the real data export). Only applied when the
                    # source explicitly declares filename keywords, and only
                    # when exactly one candidate's filename matches.
                    keyword_matches = [
                        candidate
                        for candidate in candidates
                        if any(keyword.lower() in candidate.file_path.stem.lower() for keyword in source.filename_keywords)
                    ]
                    if len(keyword_matches) == 1:
                        candidates = keyword_matches
                    else:
                        if not source.mandatory:
                            logger.warning(
                                "Optional source '%s' (%s) is ambiguous (multiple matching files); skipping (not required for this workflow).",
                                logical_name,
                                source.display_name,
                            )
                            continue
                        raise AmbiguousSourceError(source, candidates)
                else:
                    if not source.mandatory:
                        logger.warning(
                            "Optional source '%s' (%s) is ambiguous (multiple matching files); skipping (not required for this workflow).",
                            logical_name,
                            source.display_name,
                        )
                        continue
                    raise AmbiguousSourceError(source, candidates)

        chosen = candidates[0]
        resolved[logical_name] = ResolvedSource(
            logical_name=logical_name,
            file_path=chosen.file_path,
            sheet_name=chosen.sheet_name,
            header_row_index=chosen.header_row_index,
        )
        logger.info(
            "Resolved source '%s' (%s) -> file=%s | sheet=%s",
            logical_name,
            source.display_name,
            chosen.file_path.name,
            chosen.sheet_name,
        )

    return resolved
