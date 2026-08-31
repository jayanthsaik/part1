from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import AppConfig
from modules.derived_inventory import DerivedInventoryResult, build_ups_inventory
from modules.loader import WorkbookData
from modules.lookup_key_builder import LookupKeyResult, build_lookup_keys
from modules.moq_validator import MoqValidationResult, build_moq_validation
from modules.utils import coerce_numeric_column, normalize_ndc_key
from modules.validator import DataValidationError


@dataclass(frozen=True)
class Phase2Result:
    """All derived Phase 2 outputs."""

    ups_inventory_df: pd.DataFrame
    lookup_keys_df: pd.DataFrame
    moq_validation_df: pd.DataFrame
    statistics_df: pd.DataFrame
    output_path: Path
    upload_adjustments_df: Optional[pd.DataFrame] = None


def _normalize_header_name(value: object) -> str:
    """Normalize a column name for tolerant matching.

    Mirrors the normalization used by source discovery so that a file which
    was successfully DISCOVERED can always be PARSED: collapses case,
    leading/trailing/internal whitespace runs, and ``_``/``-`` separators.
    """
    text = str(value).replace("_", " ").replace("-", " ")
    return " ".join(text.split()).casefold()


def _prompt_quantity_source(logger) -> str:
    """Ask which quantity source to use when BOTH are present.

    Raises ``DataValidationError`` when no interactive console is available so
    an unattended run (scheduled task / double-clicked EXE) can never silently
    pick the wrong source and produce incorrect UPS Inventory values.
    """
    if not sys.stdin or not sys.stdin.isatty():
        raise DataValidationError(
            "Both a Morning Completed Order Book and an Upload Sheet were found "
            "in the input folder, but this run is not interactive so the correct "
            "source cannot be confirmed.\n\n"
            "Remove ONE of the two files from the input folder and re-run."
        )

    print("\nBoth quantity sources were found in the input folder:")
    print("  1. Morning Completed Order Book  (pre-processed, no filtering)")
    print("  2. Upload Sheet                  (filtered on Reason Code / Action)")
    while True:
        answer = input("Which should be used for Sales Order Qty? [1/2]: ").strip()
        if answer == "1":
            logger.info("User selected Morning Completed Order Book.")
            return "morning"
        if answer == "2":
            logger.info("User selected Upload Sheet.")
            return "upload"
        print("Please enter 1 or 2.")


def _build_quantity_adjustments(
    source_df: pd.DataFrame,
    source_name: str,
    apply_filters: bool,
    logger,
) -> pd.DataFrame:
    """Aggregate Sales Order Qty by NDC from the chosen quantity source.

    ONLY "NDC Code" and "Sales Order Qty" are read. For the Morning Completed
    Order Book this is deliberate: its own "UPS Inventory" column is IGNORED
    and UPS Inventory is always recomputed downstream as
    MAX(0, Inventory - Total - Upload_Qty), so the value can never be
    double-counted.

    ``apply_filters=True``  -> Upload Sheet: keep only rows where Reason Code
    is 1 or 4 AND Action starts with "Y" (case-insensitive), e.g. "Y",
    "Yes", "Y-Pricing issue", "Y-MOQ Issue ask cust to adjust...".

    ``apply_filters=False`` -> Morning Completed Order Book: already processed,
    so every row counts.
    """
    columns = {_normalize_header_name(column): column for column in source_df.columns}
    ndc_column = columns.get("ndc code") or columns.get("ndc")
    qty_column = columns.get("sales order qty")

    if ndc_column is None or qty_column is None:
        raise DataValidationError(
            f"{source_name} is missing required columns "
            "'NDC Code' and/or 'Sales Order Qty'."
        )

    working = source_df.copy()
    total_rows = len(working)

    if apply_filters:
        reason_column = columns.get("reason code")
        action_column = columns.get("action")
        if reason_column is None or action_column is None:
            raise DataValidationError(
                f"{source_name} is missing 'Reason Code' and/or 'Action', which "
                "are required to filter the upload sheet."
            )
        # Reason Code may arrive as 1, "1", or 1.0 (Excel float coercion).
        reason_values = working[reason_column].map(_normalize_reason_code)
        action_values = working[action_column].astype("string").str.strip().str.casefold()
        # Any value starting with "Y" counts as approved, including annotated
        # forms like "Y-Pricing issue" or "Y-MOQ Issue ask cust to adjust...",
        # not just a bare "Y"/"Yes".
        action_is_yes = action_values.str.startswith("y", na=False)
        working = working[reason_values.isin({"1", "4"}) & action_is_yes].copy()

    working[ndc_column] = working[ndc_column].map(
        lambda value: normalize_ndc_key(value) if pd.notna(value) else value
    )
    working[qty_column] = coerce_numeric_column(working[qty_column]).fillna(0)

    adjustments = (
        working.groupby(ndc_column, dropna=True)[qty_column]
        .sum()
        .rename("Upload_Qty")
        .reset_index()
        .rename(columns={ndc_column: "NDC"})
    )

    logger.info(
        "%s adjustments | source_rows=%s | rows_used=%s | ndc_groups=%s",
        source_name,
        total_rows,
        len(working),
        int(adjustments["NDC"].nunique(dropna=True)),
    )
    return adjustments


def _normalize_reason_code(value: object) -> str:
    """Return a Reason Code as a comparable string ("1", "4", ...).

    Tolerates numeric, float ("1.0") and text representations.
    """
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def run_phase2(loaded_workbooks: dict[str, WorkbookData], config: AppConfig, logger, execution_time_seconds: float) -> Phase2Result:
    """Generate all Phase 2 derived datasets from Phase 1 loaded workbooks."""
    inventory_df = _get_sheet_dataframe(loaded_workbooks, "inventory")
    open_orders_df = _get_sheet_dataframe(loaded_workbooks, "open_order_summary")

    # BUSINESS RULE: the Sales Order Qty used to adjust UPS Inventory comes
    # from the Morning Completed Order Book when supplied, otherwise from the
    # Upload Sheet. At least ONE of the two MUST be present. Both sources are
    # resolved by header signature via the source registry -- never by
    # filename -- so the client may name either file anything.
    morning_df = _get_optional_sheet_dataframe(loaded_workbooks, "morning_completed_orderbook")
    upload_df = _get_optional_sheet_dataframe(loaded_workbooks, "upload_sheet")

    if morning_df is None and upload_df is None:
        raise DataValidationError(
            "No quantity source found.\n\n"
            "This application requires EITHER a Morning Completed Order Book "
            "OR an Upload Sheet in the input folder.\n\n"
            "Place one of these files in the input folder and re-run."
        )

    if morning_df is not None and upload_df is not None:
        choice = _prompt_quantity_source(logger)
        if choice == "morning":
            adjustment_df, adjustment_source_name, apply_filters = morning_df, "Morning Completed Order Book", False
        else:
            adjustment_df, adjustment_source_name, apply_filters = upload_df, "Upload Sheet", True
    elif morning_df is not None:
        # Already processed -- no Reason Code / Action filtering.
        adjustment_df, adjustment_source_name, apply_filters = morning_df, "Morning Completed Order Book", False
    else:
        adjustment_df, adjustment_source_name, apply_filters = upload_df, "Upload Sheet", True

    logger.info(
        "Sales Order Qty adjustment source: %s (row filtering=%s)",
        adjustment_source_name,
        "yes" if apply_filters else "no (pre-processed)",
    )

    upload_adjustments_df = _build_quantity_adjustments(
        adjustment_df,
        adjustment_source_name,
        apply_filters,
        logger,
    )

    # BUSINESS RULE: quantity adjustments are netted out of UPS Inventory
    # BEFORE the floor-to-zero step (see derived_inventory.build_ups_inventory).
    inventory_result = build_ups_inventory(
        inventory_df,
        open_orders_df,
        config.phase2,
        logger,
        upload_adjustments_df=upload_adjustments_df,
    )

    # Lookup Key and MOQ Validation derived datasets were historically built
    # from an optional "sales_summary" INPUT. That source has been removed:
    # Sales Summary is generated by this pipeline as an OUTPUT, and both
    # derived datasets are now superseded downstream -- Phase 3 builds the
    # canonical Lookup itself (see master_builder._ensure_lookup) and the MOQ
    # rule is computed directly from the Orderbook's own Sales Order Qty /
    # Pack Size (MOQ) columns (see business_rules / report_formatter). They are
    # retained as empty, correctly-shaped results purely so the Phase 2
    # statistics sheet and the DEBUG-only derived workbook keep their schema.
    lookup_result = LookupKeyResult(dataframe=pd.DataFrame(), total_rows=0, values_normalized=0, duplicate_lookup_keys=0)
    moq_result = MoqValidationResult(dataframe=pd.DataFrame(), failure_count=0)

    statistics_df = _build_statistics(inventory_result, lookup_result, moq_result, execution_time_seconds)

    output_path = config.output_dir / config.phase2.derived_workbook_name
    # DEBUG-ONLY ARTIFACT. Every derived dataset above is returned in memory
    # via Phase2Result and consumed directly by Phase 3/4, so skipping this
    # write cannot change any business value -- it is a pure diagnostic
    # side effect. In production only POB.xlsx is written.
    if getattr(config, "debug_mode", False):
        _write_derived_workbook(
            inventory_result.dataframe,
            lookup_result.dataframe,
            moq_result.dataframe,
            statistics_df,
            output_path,
        )
        logger.info("[DEBUG] Wrote Phase 2 derived workbook to %s", output_path)
    else:
        logger.info("Phase 2 derived datasets kept in memory (production mode; no intermediate workbook written)")

    return Phase2Result(
        ups_inventory_df=inventory_result.dataframe,
        lookup_keys_df=lookup_result.dataframe,
        moq_validation_df=moq_result.dataframe,
        statistics_df=statistics_df,
        output_path=output_path,
        upload_adjustments_df=upload_adjustments_df,
    )


def _get_sheet_dataframe(loaded_workbooks: dict[str, WorkbookData], workbook_key: str) -> pd.DataFrame:
    """Return a copy of the first worksheet dataframe for a loaded workbook key."""
    workbook_data = loaded_workbooks.get(workbook_key)
    if workbook_data is None:
        raise ValueError(f"Required workbook '{workbook_key}' not found in loaded workbook set")
    if not workbook_data.worksheets:
        raise ValueError(f"Workbook '{workbook_key}' does not contain any worksheets")
    first_sheet = next(iter(workbook_data.worksheets.values()))
    return first_sheet.dataframe.copy()


def _get_optional_sheet_dataframe(loaded_workbooks: dict[str, WorkbookData], workbook_key: str) -> pd.DataFrame | None:
    """Return a copy of the first worksheet dataframe for an optional workbook key, or None if absent.

    Unlike ``_get_sheet_dataframe``, this never raises when the workbook is
    missing -- used for logical sources marked ``mandatory=False`` in
    ``source_registry.py`` (e.g. "sales_summary", which this pipeline
    actually generates as an output rather than receiving as an input).
    """
    workbook_data = loaded_workbooks.get(workbook_key)
    if workbook_data is None or not workbook_data.worksheets:
        return None
    first_sheet = next(iter(workbook_data.worksheets.values()))
    return first_sheet.dataframe.copy()


def _build_statistics(
    inventory_result: DerivedInventoryResult,
    lookup_result: LookupKeyResult,
    moq_result: MoqValidationResult,
    execution_time_seconds: float,
) -> pd.DataFrame:
    """Build the required Phase 2 statistics sheet."""
    return pd.DataFrame(
        [
            {
                "Inventory Rows": inventory_result.inventory_rows,
                "Allocated Rows": inventory_result.allocated_rows,
                "Unique NDCs": inventory_result.unique_ndcs,
                "Missing Inventory": inventory_result.missing_inventory,
                "Missing Allocations": inventory_result.missing_allocations,
                "MOQ Failures": moq_result.failure_count,
                "Execution Time": round(execution_time_seconds, 2),
            }
        ]
    )


def _write_derived_workbook(
    ups_inventory_df: pd.DataFrame,
    lookup_keys_df: pd.DataFrame,
    moq_validation_df: pd.DataFrame,
    statistics_df: pd.DataFrame,
    output_path: Path,
) -> None:
    """Write the Phase 2 derived datasets workbook with basic readability formatting."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        ups_inventory_df.to_excel(writer, sheet_name="UPS_Inventory", index=False)
        lookup_keys_df.to_excel(writer, sheet_name="Lookup_Keys", index=False)
        moq_validation_df.to_excel(writer, sheet_name="MOQ_Validation", index=False)
        statistics_df.to_excel(writer, sheet_name="Statistics", index=False)

    workbook = load_workbook(output_path)
    for worksheet in workbook.worksheets:
        _format_sheet(worksheet)
    workbook.save(output_path)


def _format_sheet(worksheet) -> None:
    """Apply basic workbook readability formatting to a worksheet."""
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for column_cells in worksheet.columns:
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 55)