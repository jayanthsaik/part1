"""Tests for the low UPS Inventory yellow highlight on the aggregated
Summary sheets (Sales_Summary.xlsx and the POB.xlsx "Summary" sheet).

Business rule under test:

    On the Summary sheets, when "Max of UPS Inventory" < 10,000 that
    single cell is filled yellow. Only that cell is filled -- never the
    whole row, and never any other column.

This rule was MOVED off the client-facing POB.xlsx "Orderbook" sheet's
"UPS Inventory" column (see test_moq_row_highlight.py, which pins that
the Orderbook sheet no longer highlights it).
"""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from modules.business_rules import LOW_UPS_INVENTORY_THRESHOLD  # noqa: E402
from modules.report_formatter import (  # noqa: E402
    FILL_LOW_INVENTORY_YELLOW,
    apply_low_ups_inventory_formatting,
    write_dataframe_sheet,
)

YELLOW = FILL_LOW_INVENTORY_YELLOW.fgColor.rgb


def _fill_rgb(cell):
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return None
    return fill.fgColor.rgb


def _build_sheet(values: list) -> tuple:
    workbook = Workbook()
    df = pd.DataFrame(
        {
            "Material Description": [f"Product {i}" for i in range(len(values))],
            "Sold-to party Name": ["Customer"] * len(values),
            "Max of UPS Inventory": values,
            "Sum of Sales Order Qty": [100] * len(values),
        }
    )
    worksheet = write_dataframe_sheet(workbook, "Sales Summary", df)
    count = apply_low_ups_inventory_formatting(worksheet)
    header = {str(c.value): c.column for c in worksheet[1]}
    return worksheet, header, count


class TestSalesSummaryLowUpsInventoryHighlight(unittest.TestCase):
    def test_below_threshold_is_highlighted(self):
        worksheet, header, count = _build_sheet([9_999])
        cell = worksheet.cell(row=2, column=header["Max of UPS Inventory"])
        self.assertEqual(_fill_rgb(cell), YELLOW)
        self.assertEqual(count, 1)

    def test_at_threshold_is_not_highlighted(self):
        worksheet, header, count = _build_sheet([LOW_UPS_INVENTORY_THRESHOLD])
        cell = worksheet.cell(row=2, column=header["Max of UPS Inventory"])
        self.assertIsNone(_fill_rgb(cell))
        self.assertEqual(count, 0)

    def test_above_threshold_is_not_highlighted(self):
        worksheet, header, count = _build_sheet([25_000])
        cell = worksheet.cell(row=2, column=header["Max of UPS Inventory"])
        self.assertIsNone(_fill_rgb(cell))
        self.assertEqual(count, 0)

    def test_zero_is_highlighted(self):
        worksheet, header, count = _build_sheet([0])
        cell = worksheet.cell(row=2, column=header["Max of UPS Inventory"])
        self.assertEqual(_fill_rgb(cell), YELLOW)
        self.assertEqual(count, 1)

    def test_only_the_ups_inventory_cell_is_filled(self):
        worksheet, header, _ = _build_sheet([500])
        for name, idx in header.items():
            if name == "Max of UPS Inventory":
                continue
            self.assertIsNone(
                _fill_rgb(worksheet.cell(row=2, column=idx)),
                f"'{name}' must not be highlighted",
            )

    def test_blank_and_non_numeric_values_are_skipped(self):
        worksheet, header, count = _build_sheet([None, "N/A"])
        column = header["Max of UPS Inventory"]
        self.assertIsNone(_fill_rgb(worksheet.cell(row=2, column=column)))
        self.assertIsNone(_fill_rgb(worksheet.cell(row=3, column=column)))
        self.assertEqual(count, 0)

    def test_only_qualifying_rows_are_highlighted(self):
        worksheet, header, count = _build_sheet([100, 50_000, 9_999, 10_001])
        column = header["Max of UPS Inventory"]
        self.assertEqual(_fill_rgb(worksheet.cell(row=2, column=column)), YELLOW)
        self.assertIsNone(_fill_rgb(worksheet.cell(row=3, column=column)))
        self.assertEqual(_fill_rgb(worksheet.cell(row=4, column=column)), YELLOW)
        self.assertIsNone(_fill_rgb(worksheet.cell(row=5, column=column)))
        self.assertEqual(count, 2)

    def test_missing_column_is_a_no_op(self):
        workbook = Workbook()
        df = pd.DataFrame({"Material Description": ["Product"], "Sum of Sales Order Qty": [10]})
        worksheet = write_dataframe_sheet(workbook, "Sales Summary", df)
        self.assertEqual(apply_low_ups_inventory_formatting(worksheet), 0)

    def test_column_is_located_by_header_not_position(self):
        # Reordering columns must not change which cells get highlighted.
        workbook = Workbook()
        df = pd.DataFrame(
            {
                "Max of UPS Inventory": [500],
                "Material Description": ["Product"],
                "Sum of Sales Order Qty": [10],
            }
        )
        worksheet = write_dataframe_sheet(workbook, "Sales Summary", df)
        apply_low_ups_inventory_formatting(worksheet)
        header = {str(c.value): c.column for c in worksheet[1]}
        self.assertEqual(_fill_rgb(worksheet.cell(row=2, column=header["Max of UPS Inventory"])), YELLOW)
        self.assertIsNone(_fill_rgb(worksheet.cell(row=2, column=header["Material Description"])))


class TestPobSummaryLowUpsInventoryHighlight(unittest.TestCase):
    """The same rule must also apply to the client-facing POB.xlsx
    "Summary" sheet, which carries the identical "Max of UPS Inventory"
    header (see ``_summary_reference_columns`` in phase4_manager.py)."""

    def _pob_summary_sheet(self, values: list):
        workbook = Workbook()
        # Exact reference Summary header order, incl. the leading space in
        # " Material Description" and dynamic month labels.
        df = pd.DataFrame(
            {
                " Material Description": [f"Product {i}" for i in range(len(values))],
                "Sold-to party Name": ["Customer"] * len(values),
                "Lookup": [f"KEY{i}" for i in range(len(values))],
                "NDC Code": ["64380016101"] * len(values),
                "Max of UPS Inventory": values,
                "Sum of Sales Order Qty": [100] * len(values),
                "Max of Sales Qty MTD": [5] * len(values),
                "Max of Forecast Qty": [7] * len(values),
                "John": [""] * len(values),
                "Feb": [1] * len(values),
                "Mar": [2] * len(values),
                "Apr": [3] * len(values),
                "Avg": [2] * len(values),
                "Buying Group": [""] * len(values),
                "Award Type": [""] * len(values),
                "SC Comments": [""] * len(values),
            }
        )
        worksheet = write_dataframe_sheet(workbook, "Summary", df)
        count = apply_low_ups_inventory_formatting(worksheet)
        header = {str(c.value): c.column for c in worksheet[1]}
        return worksheet, header, count

    def test_below_threshold_is_highlighted_on_pob_summary(self):
        worksheet, header, count = self._pob_summary_sheet([9_999])
        self.assertEqual(_fill_rgb(worksheet.cell(row=2, column=header["Max of UPS Inventory"])), YELLOW)
        self.assertEqual(count, 1)

    def test_above_threshold_is_not_highlighted_on_pob_summary(self):
        worksheet, header, count = self._pob_summary_sheet([10_000])
        self.assertIsNone(_fill_rgb(worksheet.cell(row=2, column=header["Max of UPS Inventory"])))
        self.assertEqual(count, 0)

    def test_no_other_reference_column_is_filled(self):
        worksheet, header, _ = self._pob_summary_sheet([250])
        for name, idx in header.items():
            if name == "Max of UPS Inventory":
                continue
            self.assertIsNone(
                _fill_rgb(worksheet.cell(row=2, column=idx)),
                f"'{name}' must not be highlighted on the POB Summary sheet",
            )


if __name__ == "__main__":
    unittest.main()
