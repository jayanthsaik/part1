from __future__ import annotations

import logging
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from modules.report_formatter import (  # noqa: E402
    FILL_CANCEL_RED,
    FILL_HOLD_BLUE,
    FILL_LOW_INVENTORY_YELLOW,
    apply_low_ups_inventory_formatting,
    apply_summary_action_formatting,
    write_dataframe_sheet,
)
from modules.sales_summary_builder import (  # noqa: E402
    SUMMARY_ACTION_FLAG_COLUMN,
    build_sales_summary_aggregation,
)

RED = FILL_CANCEL_RED.start_color.rgb
BLUE = FILL_HOLD_BLUE.start_color.rgb
YELLOW = FILL_LOW_INVENTORY_YELLOW.start_color.rgb
NO_FILL = "00000000"


@pytest.fixture()
def logger():
    return logging.getLogger("test_summary_action_row_colors")


def _detail_row(lookup: str, action, qty: int = 10, ups: int = 5_000):
    return {
        "Material Description": f"Mat {lookup}",
        "Sold-to party Name": f"Cust {lookup}",
        "Lookup": lookup,
        "NDC Code": lookup,
        "UPS Inventory": ups,
        "Sales Order Qty": qty,
        "Sales Qty MTD": 1,
        "Forecast Qty": 1,
        "Action": action,
    }


def _fill_of(worksheet, row_number: int) -> str:
    return worksheet.cell(row=row_number, column=1).fill.start_color.rgb


def _write(summary_df: pd.DataFrame):
    workbook = Workbook()
    worksheet = write_dataframe_sheet(workbook, "Summary", summary_df)
    return worksheet


def test_cancel_takes_precedence_over_hold_within_one_group(logger):
    """A Lookup containing BOTH a Cancel and a Hold line must roll up to
    Cancel, matching the Orderbook sheet's Cancel-first precedence."""
    detail = pd.DataFrame(
        [
            _detail_row("A", "Cancel"),
            _detail_row("A", "Hold"),
            _detail_row("A", None),
        ]
    )

    result = build_sales_summary_aggregation(detail, logger)

    assert len(result.dataframe) == 1
    assert result.dataframe[SUMMARY_ACTION_FLAG_COLUMN].iloc[0] == "Cancel"


def test_hold_applies_only_when_no_cancel_present(logger):
    detail = pd.DataFrame([_detail_row("B", "Hold"), _detail_row("B", "")])

    result = build_sales_summary_aggregation(detail, logger)

    assert result.dataframe[SUMMARY_ACTION_FLAG_COLUMN].iloc[0] == "Hold"


def test_rows_are_colored_red_blue_and_unfilled(logger):
    detail = pd.DataFrame(
        [
            _detail_row("A", "Cancel"),
            _detail_row("B", "Hold"),
            _detail_row("C", None),
        ]
    )
    summary_df = build_sales_summary_aggregation(detail, logger).dataframe
    worksheet = _write(summary_df)

    cancel_rows, hold_rows = apply_summary_action_formatting(worksheet, summary_df)

    assert (cancel_rows, hold_rows) == (1, 1)
    # Rows are sorted by the groupby keys -> A, B, C on worksheet rows 2,3,4.
    assert _fill_of(worksheet, 2) == RED
    assert _fill_of(worksheet, 3) == BLUE
    assert _fill_of(worksheet, 4) == NO_FILL


def test_action_matching_is_case_and_whitespace_insensitive(logger):
    detail = pd.DataFrame([_detail_row("A", "  cANCel "), _detail_row("B", "HOLD")])
    summary_df = build_sales_summary_aggregation(detail, logger).dataframe

    assert summary_df[SUMMARY_ACTION_FLAG_COLUMN].tolist() == ["Cancel", "Hold"]


def test_whole_row_is_filled_not_just_first_cell(logger):
    detail = pd.DataFrame([_detail_row("A", "Cancel")])
    summary_df = build_sales_summary_aggregation(detail, logger).dataframe
    worksheet = _write(summary_df)

    apply_summary_action_formatting(worksheet, summary_df)

    for column_index in range(1, worksheet.max_column + 1):
        assert worksheet.cell(row=2, column=column_index).fill.start_color.rgb == RED


def test_action_fill_overrides_low_ups_inventory_yellow(logger):
    """Cancel/Hold is applied last and must win over the yellow
    'Max of UPS Inventory' cell highlight."""
    detail = pd.DataFrame([_detail_row("A", "Cancel", ups=1)])
    summary_df = build_sales_summary_aggregation(detail, logger).dataframe
    worksheet = _write(summary_df)

    highlighted = apply_low_ups_inventory_formatting(worksheet)
    assert highlighted == 1  # yellow applied first

    apply_summary_action_formatting(worksheet, summary_df)

    ups_index = [c.value for c in worksheet[1]].index("Max of UPS Inventory") + 1
    assert worksheet.cell(row=2, column=ups_index).fill.start_color.rgb == RED


def test_missing_action_column_is_handled_safely(logger):
    detail = pd.DataFrame([_detail_row("A", "Cancel")]).drop(columns=["Action"])

    summary_df = build_sales_summary_aggregation(detail, logger).dataframe
    worksheet = _write(summary_df)

    assert summary_df[SUMMARY_ACTION_FLAG_COLUMN].isna().all()
    assert apply_summary_action_formatting(worksheet, summary_df) == (0, 0)
    assert _fill_of(worksheet, 2) == NO_FILL


def test_missing_flag_column_is_a_no_op(logger):
    summary_df = pd.DataFrame({"Lookup": ["A"], "Max of UPS Inventory": [5]})
    worksheet = _write(summary_df)

    assert apply_summary_action_formatting(worksheet, summary_df) == (0, 0)


def test_aggregation_values_are_unchanged_by_the_new_flag(logger):
    """The Action rollup must not disturb any existing aggregated value."""
    detail = pd.DataFrame(
        [
            _detail_row("A", "Cancel", qty=10, ups=100),
            _detail_row("A", "Hold", qty=5, ups=300),
        ]
    )

    summary_df = build_sales_summary_aggregation(detail, logger).dataframe

    assert summary_df["Sum of Sales Order Qty"].iloc[0] == 15
    assert summary_df["Max of UPS Inventory"].iloc[0] == 300


def test_flag_never_leaks_into_client_facing_summary_sheet(logger):
    """The internal flag must be stripped by the reference projection."""
    from modules.phase4_manager import _to_reference_summary_dataframe

    detail = pd.DataFrame([_detail_row("A", "Cancel")])
    summary_df = build_sales_summary_aggregation(detail, logger).dataframe
    summary_df["Average"] = 0

    reference_df = _to_reference_summary_dataframe(summary_df, ("Apr", "May", "Jun"))

    assert SUMMARY_ACTION_FLAG_COLUMN not in reference_df.columns
