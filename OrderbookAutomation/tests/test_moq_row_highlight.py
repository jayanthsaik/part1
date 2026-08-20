"""Tests for the MOQ Issue SINGLE-CELL yellow highlight in the final
Orderbook worksheet (modules/report_formatter.py), per the documented
manual process:

"...divide Sales Order Quantity by Pack Size (MOQ). Any lines with a
number after the decimal point are not in MOQ."

Business rule under test:

- Sales Order Qty / Pack Size (MOQ) not an exact multiple -> MOQ Issue.
- ONLY the "Sales Order Qty" cell is highlighted yellow (never the whole
  row). This follows the same "cell overrides row" pattern already used
  for the orange Price Issue cells.
- The MOQ Issue flag is computed independently from the Sales Order Qty
  and Pack Size (MOQ) worksheet cells, so the fill is stable and does not
  depend on any later UPS Inventory value.
- Blank/zero/non-numeric Pack Size or Sales Order Qty -> no highlight,
  no error.
- Independent of the Low_UPS_Inventory rule (which lives on the Summary
  sheet's "Max of UPS Inventory" column, not on POB's Orderbook sheet).
- Controlled Product pink fills the whole row, but the yellow MOQ cell
  remains visible on top of it.
- Cancel (red) and Hold (blue) take FULL precedence and overwrite the
  yellow MOQ cell.
"""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from modules.report_formatter import (  # noqa: E402
    FILL_CANCEL_RED,
    FILL_CONTROLLED_PRODUCT_PINK,
    FILL_HOLD_BLUE,
    FILL_LOW_INVENTORY_YELLOW,
    _is_moq_issue,
    apply_business_rule_formatting,
    write_dataframe_sheet,
)

YELLOW = FILL_LOW_INVENTORY_YELLOW.fgColor.rgb
PINK = FILL_CONTROLLED_PRODUCT_PINK.fgColor.rgb
RED = FILL_CANCEL_RED.fgColor.rgb
BLUE = FILL_HOLD_BLUE.fgColor.rgb

MOQ_CELL = "Sales Order Qty"


def _fill_rgb(cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return None
    return str(fill.fgColor.rgb) if fill.fgColor is not None else None


def _row(sales_order_qty, pack_size, ups_inventory=50000, action=""):
    return {
        "Sales Order No.": "SO1",
        "Item No.": "1",
        "NDC Code": "123456789",
        "Material Description": "Widget",
        "UPS Inventory": ups_inventory,
        "Sales Order Qty": sales_order_qty,
        "Pack Size (MOQ)": pack_size,
        "Action": action,
        "Unit Price": 10,
        "WAC/BG price in EDI": 10,
    }


def _build_sheet(rows: list[dict], low_ups_inventory_flags=None, controlled_flags=None):
    workbook = Workbook()
    df = pd.DataFrame(rows)
    worksheet = write_dataframe_sheet(workbook, "Orderbook", df)
    flags_df = df.copy()
    flags_df["Controlled_Product"] = controlled_flags if controlled_flags is not None else [False] * len(rows)
    flags_df["Price_Issue"] = False
    flags_df["Low_UPS_Inventory"] = low_ups_inventory_flags if low_ups_inventory_flags is not None else [False] * len(rows)
    apply_business_rule_formatting(worksheet, flags_df)
    return worksheet, df


def _header(worksheet) -> dict[str, int]:
    return {str(c.value): c.column for c in worksheet[1]}


class TestIsMoqIssueUnit(unittest.TestCase):
    def test_48_over_24_no_issue(self):
        self.assertFalse(_is_moq_issue(48, 24))

    def test_50_over_24_is_issue(self):
        self.assertTrue(_is_moq_issue(50, 24))

    def test_72_over_24_no_issue(self):
        self.assertFalse(_is_moq_issue(72, 24))

    def test_75_over_24_is_issue(self):
        self.assertTrue(_is_moq_issue(75, 24))

    def test_pack_size_blank_no_issue(self):
        self.assertFalse(_is_moq_issue(50, None))

    def test_pack_size_zero_no_issue(self):
        self.assertFalse(_is_moq_issue(50, 0))

    def test_pack_size_non_numeric_no_issue(self):
        self.assertFalse(_is_moq_issue(50, "xyz"))


class TestMoqCellHighlight(unittest.TestCase):
    """The MOQ rule fills ONLY the "Sales Order Qty" cell."""

    def _assert_only_moq_cell_yellow(self, worksheet, row_number=2):
        header = _header(worksheet)
        for name, idx in header.items():
            rgb = _fill_rgb(worksheet.cell(row=row_number, column=idx))
            if name == MOQ_CELL:
                self.assertEqual(rgb, YELLOW, f"'{name}' should be yellow")
            else:
                self.assertIsNone(rgb, f"'{name}' should NOT be filled (row must not be whole-row yellow)")

    def _assert_no_fill_anywhere(self, worksheet, row_number=2):
        header = _header(worksheet)
        for name, idx in header.items():
            self.assertIsNone(
                _fill_rgb(worksheet.cell(row=row_number, column=idx)),
                f"'{name}' should not be filled",
            )

    def test_48_over_24_no_yellow(self):
        worksheet, _ = _build_sheet([_row(48, 24)])
        self._assert_no_fill_anywhere(worksheet)

    def test_50_over_24_only_sales_order_qty_cell_yellow(self):
        worksheet, _ = _build_sheet([_row(50, 24)])
        self._assert_only_moq_cell_yellow(worksheet)

    def test_72_over_24_no_yellow(self):
        worksheet, _ = _build_sheet([_row(72, 24)])
        self._assert_no_fill_anywhere(worksheet)

    def test_75_over_24_only_sales_order_qty_cell_yellow(self):
        worksheet, _ = _build_sheet([_row(75, 24)])
        self._assert_only_moq_cell_yellow(worksheet)

    def test_moq_rule_never_fills_whole_row(self):
        # Explicit regression guard for the old whole-row behaviour.
        worksheet, _ = _build_sheet([_row(75, 24)])
        header = _header(worksheet)
        yellow_cells = [name for name, idx in header.items() if _fill_rgb(worksheet.cell(row=2, column=idx)) == YELLOW]
        self.assertEqual(yellow_cells, [MOQ_CELL])

    def test_blank_pack_size_no_yellow(self):
        worksheet, _ = _build_sheet([_row(50, None)])
        self._assert_no_fill_anywhere(worksheet)

    def test_zero_pack_size_no_yellow(self):
        worksheet, _ = _build_sheet([_row(50, 0)])
        self._assert_no_fill_anywhere(worksheet)

    def test_non_numeric_pack_size_no_yellow(self):
        worksheet, _ = _build_sheet([_row(50, "xyz")])
        self._assert_no_fill_anywhere(worksheet)

    def test_ups_inventory_cell_never_yellow_from_moq_rule(self):
        # The MOQ rule must not bleed onto the UPS Inventory cell.
        worksheet, _ = _build_sheet([_row(75, 24, ups_inventory=99999)])
        header = _header(worksheet)
        self.assertIsNone(_fill_rgb(worksheet.cell(row=2, column=header["UPS Inventory"])))
        self.assertEqual(_fill_rgb(worksheet.cell(row=2, column=header[MOQ_CELL])), YELLOW)

    def test_reapplying_formatting_is_idempotent(self):
        worksheet, df = _build_sheet([_row(75, 24)])
        flags_df = df.copy()
        flags_df["Controlled_Product"] = False
        flags_df["Price_Issue"] = False
        flags_df["Low_UPS_Inventory"] = False
        apply_business_rule_formatting(worksheet, flags_df)  # re-apply
        self._assert_only_moq_cell_yellow(worksheet)

    def test_low_ups_inventory_is_not_highlighted_on_pob(self):
        # BUSINESS RULE: the low UPS Inventory yellow highlight lives on the
        # Summary sheet's "Max of UPS Inventory" column. Even when the
        # Low_UPS_Inventory flag is True, POB.xlsx must leave every cell
        # unfilled when there is no MOQ issue (48/24 is exact).
        worksheet, _ = _build_sheet([_row(48, 24)], low_ups_inventory_flags=[True])
        self._assert_no_fill_anywhere(worksheet)

    def test_moq_cell_still_yellow_when_low_ups_inventory_flag_set(self):
        worksheet, _ = _build_sheet([_row(75, 24)], low_ups_inventory_flags=[True])
        self._assert_only_moq_cell_yellow(worksheet)

    def test_multiple_rows_only_moq_row_cell_is_yellow(self):
        worksheet, _ = _build_sheet([_row(48, 24), _row(50, 24), _row(72, 24)])
        header = _header(worksheet)
        moq_idx = header[MOQ_CELL]

        # Row 2 (48/24) and row 4 (72/24): nothing filled.
        for row_number in (2, 4):
            for name, idx in header.items():
                self.assertIsNone(_fill_rgb(worksheet.cell(row=row_number, column=idx)), f"row {row_number} '{name}'")

        # Row 3 (50/24): only the Sales Order Qty cell is yellow.
        for name, idx in header.items():
            rgb = _fill_rgb(worksheet.cell(row=3, column=idx))
            self.assertEqual(rgb, YELLOW if idx == moq_idx else None, f"row 3 '{name}'")


class TestMoqFillPrecedence(unittest.TestCase):
    """Fill precedence: pink row < yellow MOQ cell < red/blue Action row."""

    def test_yellow_moq_cell_overrides_controlled_product_pink(self):
        worksheet, _ = _build_sheet([_row(75, 24)], controlled_flags=[True])
        header = _header(worksheet)
        for name, idx in header.items():
            rgb = _fill_rgb(worksheet.cell(row=2, column=idx))
            if name == MOQ_CELL:
                self.assertEqual(rgb, YELLOW, "MOQ cell must override the pink row fill")
            else:
                self.assertEqual(rgb, PINK, f"'{name}' should stay pink")

    def test_controlled_product_row_without_moq_stays_fully_pink(self):
        worksheet, _ = _build_sheet([_row(48, 24)], controlled_flags=[True])
        header = _header(worksheet)
        for name, idx in header.items():
            self.assertEqual(_fill_rgb(worksheet.cell(row=2, column=idx)), PINK, name)

    def test_cancel_red_takes_full_precedence_over_moq_yellow(self):
        worksheet, _ = _build_sheet([_row(75, 24, action="Cancel")])
        header = _header(worksheet)
        for name, idx in header.items():
            self.assertEqual(_fill_rgb(worksheet.cell(row=2, column=idx)), RED, f"'{name}' should be red")

    def test_hold_blue_takes_full_precedence_over_moq_yellow(self):
        worksheet, _ = _build_sheet([_row(75, 24, action="Hold")])
        header = _header(worksheet)
        for name, idx in header.items():
            self.assertEqual(_fill_rgb(worksheet.cell(row=2, column=idx)), BLUE, f"'{name}' should be blue")

    def test_cancel_red_takes_precedence_over_pink_and_moq(self):
        worksheet, _ = _build_sheet([_row(75, 24, action="Cancel")], controlled_flags=[True])
        header = _header(worksheet)
        for name, idx in header.items():
            self.assertEqual(_fill_rgb(worksheet.cell(row=2, column=idx)), RED, f"'{name}' should be red")


if __name__ == "__main__":
    unittest.main()
