from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import openpyxl
import pandas as pd

# Resolve the data directory relative to this script's location so the
# project works on any machine, regardless of where it is checked out.
ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "DATA_DICTIONARY.md"
WORKBOOKS = sorted(ROOT.glob("*.xlsx"))


def safe_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def is_blank(value: object) -> bool:
    return safe_text(value) == ""


def header_score(row: Iterable[object]) -> int:
    values = [value for value in row if not is_blank(value)]
    if not values:
        return -10**9
    string_count = sum(isinstance(value, str) for value in values)
    date_count = sum(isinstance(value, (datetime, date)) for value in values)
    numeric_count = sum(isinstance(value, (int, float)) and not isinstance(value, bool) for value in values)
    return string_count * 3 + date_count * 2 + len(values) - numeric_count


def detect_header_row(workbook_path: Path, sheet_name: str) -> int:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name]
    best_index = 0
    best_score = -10**9
    max_row = min(worksheet.max_row or 1, 25)
    for index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_row, values_only=True)):
        score = header_score(row)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def load_sheet_info(workbook_path: Path, sheet_name: str) -> tuple[list[str], pd.DataFrame] | None:
    header_row = detect_header_row(workbook_path, sheet_name)
    dataframe = pd.read_excel(workbook_path, sheet_name=sheet_name, header=header_row)
    dataframe = dataframe.loc[:, ~dataframe.columns.duplicated()].copy()
    headers = [str(column) for column in dataframe.columns]
    if not headers:
        return None
    return headers, dataframe


def infer_dtype(series: pd.Series, column_name: str) -> str:
    values = [value for value in series.tolist() if not is_blank(value)]
    if not values:
        return "Text"

    cleaned = [safe_text(value) for value in values]
    lower_name = safe_text(column_name).lower()

    bool_values = {"y", "n", "yes", "no", "true", "false", "0", "1"}
    if all(item.lower() in bool_values for item in cleaned):
        return "Boolean"

    numeric_values: list[float] = []
    for value in values:
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            numeric_values.append(float(value))
            continue
        if isinstance(value, str):
            cleaned_value = value.replace(",", "").replace("(", "-").replace(")", "")
            try:
                numeric_values.append(float(cleaned_value))
            except ValueError:
                pass

    if len(numeric_values) == len(values):
        return "Integer" if all(item.is_integer() for item in numeric_values) else "Decimal"

    if any(token in lower_name for token in ["date", "eta", "expiry", "expiration", "req"]):
        parsed_dates = pd.to_datetime(pd.Series(cleaned), errors="coerce")
        if parsed_dates.notna().mean() >= 0.7:
            return "Date"

    return "Text"


def sample_value(series: pd.Series) -> str:
    for value in series.tolist():
        if not is_blank(value):
            return safe_text(value)
    return ""


def nullable_flag(series: pd.Series) -> str:
    return "Yes" if series.isna().any() or any(is_blank(value) for value in series.tolist()) else "No"


def duplicate_headers(headers: list[str]) -> str:
    counts = Counter([header for header in headers if safe_text(header)])
    duplicates = [header for header, count in counts.items() if count > 1]
    return ", ".join(duplicates)


def recommended_join_keys(workbook_name: str, headers: list[str]) -> str:
    header_set = {safe_text(header).lower() for header in headers if safe_text(header)}
    name = workbook_name.lower()
    keys: list[str] = []

    def add(*items: str) -> None:
        for item in items:
            if item not in keys:
                keys.append(item)

    if name == "raw_ob.xlsx":
        add("Sales Order Qty", "Sold-to party", "Ship-to party")
    elif name == "mat_desc,_moq_,_material_#.xlsx":
        add("NDC code", "HANA Material")
    elif name == "07-30_inv.xlsx":
        add("NDC", "SKU")
    elif name == "open_order_summary.xlsx":
        add("SO#", "SKU", "PH_SOLDTO_NAME")
    elif name == "sales_summ.xlsx":
        add("Lookup", "NDC Code", "Sold-to party Name")
    elif name == "strend.xlsx":
        if "ndc code" in header_set and "sold-to party name" in header_set:
            add("NDC Code", "Sold-to party Name")
        elif "ndc code" in header_set:
            add("NDC Code")
        elif "lookup" in header_set:
            add("Lookup")
    elif name == "buying_groups.xlsx":
        add("Customer")
    elif name == "awards.xlsx":
        add("Lookup", "NDC")
    elif name == "cip.xlsx":
        add("NDC")
    elif name == "pob.xlsx":
        add("Lookup", "NDC Code", "Sold-to party Name")
    elif name == "obook.xlsx":
        add("Sales Order No.", "Item No.")
    elif name == "headers.xlsx":
        add("Sales Order No.", "Item No.", "Matl.Code", "NDC Code")
    return "; ".join(keys)


def main() -> None:
    lines = [
        "# DATA_DICTIONARY",
        "",
        "Source scan of Excel workbooks under `D:\\Jayanth`.",
        "",
    ]

    all_rows: list[dict[str, str]] = []
    for workbook_path in WORKBOOKS:
        try:
            workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
        except Exception:
            continue

        for sheet_name in workbook.sheetnames:
            try:
                loaded = load_sheet_info(workbook_path, sheet_name)
            except Exception:
                continue
            if loaded is None:
                continue

            headers, dataframe = loaded
            duplicate_names = duplicate_headers(headers)
            join_keys = recommended_join_keys(workbook_path.name, headers)
            for column_name in headers:
                series = dataframe[column_name]
                all_rows.append(
                    {
                        "Workbook Name": workbook_path.name,
                        "Worksheet Name": sheet_name,
                        "Exact Column Name": column_name,
                        "Data Type (inferred)": infer_dtype(series, column_name),
                        "Sample Value": sample_value(series),
                        "Nullable": nullable_flag(series),
                        "Duplicate Column Names": duplicate_names,
                        "Recommended Join Key(s)": join_keys,
                    }
                )

    current_workbook = None
    current_sheet = None
    for row in all_rows:
        if row["Workbook Name"] != current_workbook:
            lines.extend(["", f"## {row['Workbook Name']}", ""])
            current_workbook = row["Workbook Name"]
            current_sheet = None
        if row["Worksheet Name"] != current_sheet:
            lines.extend(
                [
                    f"### {row['Worksheet Name']}",
                    "",
                    "| Workbook Name | Worksheet Name | Exact Column Name | Data Type (inferred) | Sample Value | Nullable | Duplicate Column Names | Recommended Join Key(s) |",
                    "| --- | --- | --- | --- | --- | --- | --- | --- |",
                ]
            )
            current_sheet = row["Worksheet Name"]
        lines.append(
            "| {Workbook Name} | {Worksheet Name} | {Exact Column Name} | {Data Type (inferred)} | {Sample Value} | {Nullable} | {Duplicate Column Names} | {Recommended Join Key(s)} |".format(
                **{key: str(value).replace("|", "\\|") for key, value in row.items()}
            )
        )

    OUTPUT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
