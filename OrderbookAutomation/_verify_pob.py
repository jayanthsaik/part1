"""Deployment verification helper (development tool, never shipped).

Dumps a structural + business fingerprint of a POB.xlsx so a Python-built
and an EXE-built workbook can be compared cell-for-cell, including fills.
"""

from __future__ import annotations

import hashlib
import json
import sys

from openpyxl import load_workbook


def fingerprint(path: str) -> dict:
    workbook = load_workbook(path)
    report: dict = {"sheets": workbook.sheetnames, "detail": {}}

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        headers = [str(cell.value) for cell in worksheet[1]]

        values: list = []
        fills: list = []
        for row in range(2, worksheet.max_row + 1):
            row_values = []
            row_fills = []
            for column in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=column)
                row_values.append(str(cell.value))
                fill = cell.fill
                row_fills.append(fill.fgColor.rgb if fill and fill.fill_type == "solid" else None)
            values.append(row_values)
            fills.append(row_fills)

        report["detail"][sheet_name] = {
            "headers": headers,
            "column_count": worksheet.max_column,
            "row_count": worksheet.max_row - 1,
            "values_sha256": hashlib.sha256(json.dumps(values, sort_keys=True).encode()).hexdigest(),
            "fills_sha256": hashlib.sha256(json.dumps(fills, sort_keys=True).encode()).hexdigest(),
        }

    return report


if __name__ == "__main__":
    print(json.dumps(fingerprint(sys.argv[1]), indent=2))
